# -*- coding: utf-8 -*-
import discord_user
import discord
import asyncio
import re
import requests
import email.utils
import html
from string import Template
import jinja2
import os
channels = {'actu_fi.xml':{'id':'256392899720249344','label':'Actualités FI','fic':'actu_fi.xlsx'},
            'actu_generales.xml':{'id':'369220132121083904','label':'Actualités Autres','fic':'actu_generales.xlsx'},
            'reseaux_sociaux.xml':{'id':'293416436620197889','label':'Réseaux sociaux','fic':'actu_rs.xlsx'}}


#with open(inputFile, "r") as f1:
#    last_line = f1.readlines()[-1]
rss_length=30

def extract_text(html):
    import re
    #import urllib.request
    from bs4 import BeautifulSoup

    #html = urllib.request.urlopen('http://bgr.com/2014/10/15/google-android-5-0-lollipop-release/')
    soup = BeautifulSoup(html)
    data = soup.findAll(text=True)

    def visible(element):
        if element.parent.name in ['style', 'script', '[document]', 'head', 'title']:
            return False
        elif re.match('<!--.*-->', str(element.encode('utf-8'))):
            return False
        elif not element.parent.name in ['p','title']:
            return False
        return True

    result = filter(visible, data)
    return ' '.join(list(result))

client = discord.Client()
from jinja2 import Environment, PackageLoader, select_autoescape,FileSystemLoader
env = Environment(
    loader=FileSystemLoader('./templates'),
    autoescape=select_autoescape(['html', 'xml'])
)


from openpyxl import Workbook,load_workbook


client = discord.Client()

def getmeta(t,property):
    import re
    g = re.search(r'<meta  {0,1}property="'+property+'"[^>]*content="([^"]+)"',t)
    if g:
        return html.unescape(g.groups()[0])
    else:
        return ""

@client.event
async def on_ready():
    print('Logged in as')
    print(client.user.name)
    print(client.user.id)
    print('------')

    for flux in channels.keys():

        channel = client.get_channel(channels[flux]['id'])
        try:
            wb = load_workbook(channels[flux]['fic'])
            ws = wb.active
        except:
            wb = Workbook()
            ws = wb.active
            ws.append(['ID','Auteur','Tags','Lien','Date','Titre','Site','Description','Image'])

        lastid = max([v.value for v in ws['A'][1:]]) if (len(ws['A'])>1) else 'NOPE'
        print(lastid)
        items = []
        async for msg in client.logs_from(channel,limit=rss_length):
            if msg.id == lastid:
                print('bingo!')
                break
            links = re.search(r'(https?:[^ \n><]+)',msg.content)
            tags = re.search(r'[ \n>]*#([^ \n<>]+)',msg.content)
            if tags:
                tags = ','.join(tags.groups())
            else:
                tags = ''

            if links:
                for link in links.groups():
                    r = requests.get(link)
                    try:
                        c = r.content.decode('utf8')
                    except:
                        c = r.content.decode(r.encoding)

                    #print(extract_text(c))
                    site = getmeta(c,"og:site_name")
                    titre = getmeta(c,"og:title")
                    description = getmeta(c,"og:description")

                    date = email.utils.format_datetime(msg.timestamp)
                    image = getmeta(c,'og:image') or getmeta(c,'og:image:url')

                    ws.append([msg.id,msg.author if isinstance(msg.author,str) else msg.author.name ,tags,link,msg.timestamp,titre,site,description,image])

        wb.save('html/%s' % channels[flux]['fic'])
        items = []
        for r in list(ws.iter_rows())[1:][-30:]:

            date = email.utils.format_datetime(r[4].value)
            items.append(dict(titre=r[5].value,
                         link=r[3].value,
                         date=date,
                         site=r[6].value,
                         image=r[8].value,
                         description=r[7].value))

        with open('html/%s' % flux,'wb') as f:
            f.write(env.get_template('rss.html').render(items=items,
                                                        link="https://discordapp.com/channels/254215611264008192/"+channels[flux]['id'],
                                                        label=channels[flux]['label']
                                                        ).encode('utf8'))


    await client.logout()


try:
    client.loop.run_until_complete(client.start(discord_user.username,discord_user.password))
except:
    client.loop.run_until_complete(client.logout())
